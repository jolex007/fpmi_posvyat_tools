import openpyxl
import json
import requests
from PIL import Image

GROUPS = ['301', '302', '303', '304', '305', '306', '307', '308', '311', '312', '320', '321', '322', '323', '324', '325', '326', '327', '328', '331', '332', '351', '352', '353']
group = 'russian'

wb = openpyxl.load_workbook('table.xlsx',data_only=True)
fs = wb.active
fs_count_row = fs.max_row 
fs_count_col = fs.max_column

print('Rows:', fs_count_row)
print('Columns:', fs_count_col)

letters_mapping = {
    'FL': {
        'url': 'http://192.168.88.151',
        'colors': ['000000' for _ in range(438)],
    },
    'FR': {
        'url': 'http://192.168.88.149',
        'colors': ['000000' for _ in range(774)],
    },
    'P': {
        'url': 'http://192.168.88.146',
        'colors': ['000000' for _ in range(824)],
    },
    'ML': {
        'url': 'http://192.168.88.150',
        'colors': ['000000' for _ in range(598)],
    },
    'MR': {
        'url': 'http://192.168.88.147',
        'colors': ['000000' for _ in range(576)],
    },
    'I': {
        'url': 'http://192.168.88.145',
        'colors': ['000000' for _ in range(1043)],
    }
}

def zip_leds_array(leds_array):
    new_array = []
    prev_elem = ''
    begin = -1
    for i, elem in enumerate(leds_array):
        if elem != prev_elem:
            if begin != -1:
                new_array.append(begin)
                new_array.append(i)
                new_array.append(prev_elem)
            prev_elem = elem
            begin = i
        else:
            continue
    new_array.append(begin)
    new_array.append(len(leds_array))
    new_array.append(prev_elem)
    return new_array

request_template = {
    "v": True,
    "seg": {
        "id": 0,
        "i": []
    }
}

def parse_index(cell_text):
    tmp = str(cell_text).split('_')
    return tmp[0], int(tmp[1]) - 1

def parse_color(cell_color):
    cell_color = str(cell_color)
    return cell_color[2:]

img = Image.open('flags_for_letters/{}.png'.format(group))
pixels = img.load()
width, height = img.size
print('Size of image:', width, height)

for row in range(1,fs_count_row+1):
    for column in range(1,fs_count_col+1):
        cell = fs.cell(column=column, row=row)
        bgColor = parse_color(cell.fill.bgColor.index)
        cell_text = str(cell.value)
        if cell_text != 'None' and len(cell_text) > 0:
            letter, idx = parse_index(cell.value)
            if idx >= len(letters_mapping[letter]['colors']):
                print('ERROR:', letter, idx)
            # letters_mapping[letter]['colors'][idx] = bgColor
            imgColor = pixels[column - 1, row - 1]
            r, g, b = imgColor[0], imgColor[1], imgColor[2]
            letters_mapping[letter]['colors'][idx] = f"{r:02x}{g:02x}{b:02x}".upper()

        else:
            continue

with open('flags_for_letters_json/{}.json'.format(group), 'w') as f:
    json.dump(letters_mapping, f)

for letter, params in letters_mapping.items():
    if letter in []: # debug
        continue
    if len(params['colors']) == 0:
        continue
    for start_led in range(0, len(params['colors']), 200):
        req = request_template.copy()
        # zipped_colors = zip_leds_array(params['colors'])
        zipped_colors = params['colors'][start_led:start_led + 200]
        req["seg"]["i"] = [start_led] + zipped_colors

        print(json.dumps(req))
        print('Color "{}" letter'.format(letter))
        ping_response = requests.get('{url}/json/info'.format(url=params['url']))
        print(ping_response)
        response = requests.post('{url}/json/state'.format(url=params['url']), json=req)
        print(response)